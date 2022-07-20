#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar  2 13:56:47 2021

@author: samuelsolomon
"""

# -------------------------------------------------------------------------- #
# ---------------------------- Imported Modules ---------------------------- #

# General Modules
import os
import sys
import time as timer
import numpy as np
import pandas as pd
# Module to Sort Files in Order
from natsort import natsorted
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
from openpyxl import load_workbook, Workbook
# Openpyxl Styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font

# Import Bioelectric Analysis Files
import readDataArduino as streamDataProtocol      # Functions to Handle Data from Arduino

# -------------------------------------------------------------------------- #
# ---------------------- Extract Test Data from Excel ---------------------- #

class handlingExcelFormat:        
        
    def convertToXLSX(self, inputExcelFile):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(inputExcelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return inputExcelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        newExcelFolder = os.path.dirname(inputExcelFile) + "/Excel Files/"
        os.makedirs(newExcelFolder, exist_ok = True)
        
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(inputExcelFile)
        newExcelFile = newExcelFolder + filename + "x"
        pyexcel.save_as(file_name = inputExcelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Save New Excel name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to see if csv conversion alreayd happened
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            if excelDelimiter == "fixedWidth":
                df = pd.read_fwf(inputFile)
                df.drop(index=0, inplace=True) # drop the underlines
                df.to_excel(excelFile, index=False)
                # Load the Data from the Excel File
                xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
                xlWorksheet = xlWorkbook.worksheets[testSheetNum]
            else:
                # Make Excel WorkBook
                xlWorkbook = xl.Workbook()
                xlWorksheet = xlWorkbook.active
                # Write the Data from the CSV File to the Excel WorkBook
                with open(inputFile, "r") as inputData:
                    inReader = csv.reader(inputData, delimiter = excelDelimiter)
                    with open(excelFile, 'w+', newline=''):
                        for row in inReader:
                            xlWorksheet.append(row)    
                # Save as New Excel File
                xlWorkbook.save(excelFile)
        # Else Load the Data from the Excel File
        else:
            # Load the Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheet
    
    def splitExcelSheetsToExcelFiles(self, inputFile):
        wb = load_workbook(filename=inputFile)
        
        for sheet in wb.worksheets:
            new_wb = Workbook()
            ws = new_wb.active
            for row_data in sheet.iter_rows():
                for row_cell in row_data:
                    ws[row_cell.coordinate].value = row_cell.value
        
            new_wb.save('{0}.xlsx'.format(sheet.title))
    
    def addExcelAesthetics(self, WB_worksheet):
        # Center the Data in the Cells
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)        
        for column_cells in WB_worksheet.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            WB_worksheet.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length
            
            for cell in column_cells:
                cell.alignment = align
        # Header Style
        for cell in WB_worksheet["1:1"]:
            cell.font = Font(color='00FF0000', italic=True, bold=True)
        
        return WB_worksheet
    
    
class getExcelData(handlingExcelFormat):
    
    def extractFeatureNames(self, featureLabelFile, prependedString, appendToName = ''):
        """ Extract the Feature Names from a txt File """
        # Check if File Exists
        if not os.path.exists(featureLabelFile):
            print("The following Input File Does Not Exist:", featureLabelFile)
            sys.exit()

        # Get the Data
        fullText = ''
        with open(featureLabelFile, "r", newline='\n') as inputData:
            inReader = csv.reader(inputData)
            for row in inReader:
                for featureString in row:
                    if featureString[0] != "#":
                        fullText += featureString + ","
        
        possibleFeatures = fullText.split(prependedString)
        # Extract the Features
        featureList = []
        for feature in possibleFeatures:
            feature = feature.split("[")[-1]
            feature = feature.split("]")[0]
            feature = feature.replace(" ", "")
            feature = feature.replace("\n", "")
            
            if len(feature) != 0:
                feature = feature.split(",")
                featureList.extend(feature)
                
        featureListFull = []
        for feature in featureList:
            featureListFull.append(feature + appendToName)
        
        return featureListFull
    
    def extractData(self, ExcelSheet, startDataCol = 1, endDataCol = 2, data = None):
        # If Header Exists, Skip Until You Find the Data
        for row in ExcelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                break
        
        if data == None:
            data = [ [], [[] for channel in range(endDataCol-startDataCol)] ]
        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in ExcelSheet.iter_rows(min_col=startDataCol, min_row=dataStartRow-1, max_col=endDataCol, max_row=ExcelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            data[0].append(float(dataRow[0].value))
            for dataInd in range(1, len(dataRow)):
                data[1][dataInd-1].append(float(dataRow[dataInd].value or 0))
        
        return data
    
    def getData(self, inputFile, numberOfChannels = 1, testSheetNum = 0):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet which the user can specify using 'testSheetNum' (0-indexed).
        In the Worksheet:
            Time Data must be in Column 'A' (x-Axis)
            Biolectric Data must be in Column 'B-x' (y-Axis)
        If No Data is present in one cell of a row, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            inputFile: The Path to the Excel/TXT/CSV File Containing the Biolectric Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(inputFile):
            print("The following Input File Does Not Exist:", inputFile)
            sys.exit()
            
        # Convert to TXT and CSV Files to XLSX
        if inputFile.endswith(".txt") or inputFile.endswith(".csv"):
            # Extract Filename Information
            oldFileExtension = os.path.basename(inputFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = os.path.dirname(inputFile) + "/Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, excelSheet = self.convertToExcel(inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif inputFile.endswith(".xlsx"):
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(inputFile, data_only=True, read_only=True)
            excelSheet = xlWorkbook.worksheets[testSheetNum]
        else:
            print("The Following File is Neither CSV, TXT, Nor XLSX:", inputFile)
        print("Extracting Data from the Excel File:", inputFile)
        
        # Extract Time and Current Data from the File
        data = self.extractData(excelSheet, startDataCol = 1, endDataCol = 1 + numberOfChannels)

        xlWorkbook.close()
        # Finished Data Collection: Close Workbook and Return Data to User
        print("\tFinished Collecting Biolectric Data");
        return np.array(data)
    
    def streamTrainingData(self, trainingDataExcelFolder, labelMap, numberOfChannels, readData):
        """
        Parameters
        ----------
        trainingDataExcelFolder: The Folder with ONLY the Training Data Excel Files
        """
        # Prepare for data collection
        Training_Data = []; Training_Labels = []
                        
        # For each file in the training folder
        for excelFile in list(natsorted(os.listdir(trainingDataExcelFolder))):
            # Take each excel file
            if excelFile.endswith(".xlsx") and not excelFile.startswith("~"):
                # Get Full Path to the Excel File
                trainingExcelFile = trainingDataExcelFolder + excelFile
                print("\nLoading Excel File", trainingExcelFile)
                
                # Get the feature label from the filename
                featureLabel = None
                for possibleFeatureLabelInd in range(len(labelMap)):
                    possibleFeatureLabel = labelMap[possibleFeatureLabelInd]
                    if possibleFeatureLabel.lower() in excelFile.lower():
                        featureLabel = possibleFeatureLabelInd
                        break
                # If No Label, We Dont Want to Analyze its Features
                if featureLabel == None:
                    sys.exit("No Feature Detected in File " + excelFile)
                    
                # Read in the Excel Worksheet
                WB = xl.load_workbook(trainingExcelFile, data_only=True, read_only=True)
                WB_worksheets = WB.worksheets
                
                compiledRawData = None
                # Loop through and compile all the data in the file
                for excelSheet in WB_worksheets:
                    # Compile the data in the sheet
                    compiledRawData = self.extractData(excelSheet, startDataCol = 1, endDataCol = 1 + numberOfChannels, data = compiledRawData)
                
                print("\tExtracting Features With Label", featureLabel)
                # Analyze the data
                readData.streamExcelData(compiledRawData, predictionModel = None, actionControl = None, calibrateModel = False)
                # Save the features and labels
                Training_Data.extend(readData.eogAnalysis.featureListExact.copy())
                Training_Labels.extend(len(readData.eogAnalysis.featureListExact)*[featureLabel])

        # Return Training Data and Labels
        return np.array(Training_Data), np.array(Training_Labels)

# -------------------------------------------------------------------------- #
# -------------------------- Saving Data in Excel -------------------------- #

class saveExcelData(handlingExcelFormat):
    
    def saveData(self, time, signalData, filteredData, dataHeaders, saveDataFolder, saveExcelName, sheetName = "Pulse Data"):
        # Create Output File Directory to Save Data: If None Exists
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Path to File to Save
        excelFile = saveDataFolder + saveExcelName
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excelFile):
            print("Saving the Data as New Excel Workbook")
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        else:
            print("Excel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile, read_only=False)
            WB_worksheet = WB.create_sheet(sheetName)
        
        # Get the Header for the Data
        header = ["Time (Seconds)"]
        header.extend([dataHeader.upper() + " Raw Data" for dataHeader in dataHeaders])
        if len(filteredData) != 0:
            header.extend([dataHeader.upper() + " Filtered Data" for dataHeader in dataHeaders])        
        
        maxAddToExcelSheet = 1048500  # Max Rows in a Worksheet
        # Loop through/save all the data in batches of maxAddToExcelSheet.
        for firstIndexInFile in range(0, len(time), maxAddToExcelSheet):
            startTimer = timer.time()
            # Add the header labels to this specific file.
            WB_worksheet.append(header)
                        
            # Loop through all data to be saved within this sheet in the excel file.
            for dataInd in range(firstIndexInFile, min(firstIndexInFile+maxAddToExcelSheet, len(time))):
                # Organize all the data
                row = [time[dataInd]]
                row.extend([dataCol[dataInd] for dataCol in signalData])
                if len(filteredData) != 0:
                    row.extend([dataCol[dataInd] for dataCol in filteredData])
                
                # Save the row to the worksheet
                WB_worksheet.append(row)
    
            # Add Excel Aesthetics
            WB_worksheet = self.addExcelAesthetics(WB_worksheet)  

            # Add Sheet
            WB_worksheet = WB.create_sheet(sheetName)
            
            # If I need to use another sheet
            if firstIndexInFile + maxAddToExcelSheet < len(time):
                # Keep track of how long it is taking.
                endTimer = timer.time()
                numberOfSheetsLeft = 1+(len(time) - firstIndexInFile - maxAddToExcelSheet)//maxAddToExcelSheet
                timeRemaining = (endTimer - startTimer)*numberOfSheetsLeft
                print("\tEstimated Time Remaining " + str(timeRemaining) + " seconds; Excel Sheets Left to Add: " + str(numberOfSheetsLeft))
            
        WB.remove(WB_worksheet)
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()
        
        
    
    def saveLabeledPoints(self, signalData, signalLabelsTrue, signalLabelsPredicted, saveDataFolder, saveExcelName, sheetName = "Signal Data and Labels"): 
        # Create Output File Directory to Save Data: If None
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Path to File to Save
        excelFile = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excelFile):
            print("\nSaving the Data as New Excel Workbook", "\n")
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        # Loading in Previous Excel File and Creating New Sheet
        else:
            print("Excel File Already Exists. Loading File")
            WB = xl.load_workbook(excelFile, read_only=False)
            WB_worksheet = WB.create_sheet(sheetName)
            print("Saving Sheet as", sheetName, "\n")
        
        header = ['Dimension ' + str(featureNum) for featureNum in range(1, 1+len(signalData[0]))]
        header.extend(['Signal Labels True', 'Signal Labels Predicted'])
        WB_worksheet.append(header)
        
        # Save Data to Worksheet
        for pointInd in range(len(signalData)):
            # Get labels
            row = [signalData[pointInd][featureNum] for featureNum in range(len(signalData[pointInd]))]
            row.extend([signalLabelsTrue[pointInd], signalLabelsPredicted[pointInd]])
            WB_worksheet.append(row)
        
        # Center the Data in the Cells
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)        
        for column_cells in WB_worksheet.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            WB_worksheet.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length
            
            for cell in column_cells:
                cell.alignment = align
        # Header Style
        for cell in WB_worksheet["1:1"]:
            cell.font = Font(color='00FF0000', italic=True, bold=True)
            
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()
