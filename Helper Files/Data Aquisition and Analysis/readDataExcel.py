#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar  2 13:56:47 2021

@author: samuelsolomon
"""

# General Modules
import os
import re
import sys
import numpy as np
# Read/Write to Excel
import shutil
import openpyxl as xl
# Style Excel Data
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font

# --------------------------------------------------------------------------- #
# --------------------- Extract Test Data from Excel ------------------------ #

class readExcel():
    def __init__(self, analysisProtocol):
        # Save Protocol
        self.analysisProtocol = analysisProtocol
        # Comonly Used Variables
        self.numChannels = analysisProtocol.numChannels
        self.moveDataFinger = analysisProtocol.moveDataFinger
        self.numTimePoints = analysisProtocol.numTimePoints
        
    def streamExcelData(self, testDataExcelFile, plotStreamedData = False, testSheetNum = 0, predictionModel = None, actionControl=None, analyzeSheet = None):
        """
        Extracts Biolectric Data from Excel Document (.xlsx). Data can be in any
        worksheet, which the user can specify using 'testSheetNum'.
        In the Worksheet:
            TimePoints Must be in Columns 1
            Biolectric Voltages Must Follow in Columns 2-numChannels
        If No Data is present in a cell, it will be read in as zero.
        If No TimePoint, the Loop Will Exit
        --------------------------------------------------------------------------
        Input Variable Definitions:
            testDataExcelFile: The Path to the Excel File Containing the Channel Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) of the data.
        --------------------------------------------------------------------------
        """
        # Reset Global Variable in Case it Was Previously Populated
        self.analysisProtocol.resetGlobalVariables()
    
        # Load Data from Excel File if No Sheet Given
        if analyzeSheet == None:
            print("\nAttempting to Extract Data from the Excel File:", testDataExcelFile)
            WB = xl.load_workbook(testDataExcelFile, data_only=True,read_only=True)
            WB_worksheets = WB.worksheets
            analyzeSheet = WB_worksheets[testSheetNum]
        
        # If Header Exists, Skip Until You Find the Data Data
        for row in analyzeSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row
                break

        dataFinger = 0; pointNum = 0
        # Loop Through the Excel Worksheet to collect all the data
        for arduinoData in analyzeSheet.iter_rows(min_col=1, min_row=dataStartRow, max_col=self.numChannels+1, max_row=analyzeSheet.max_row):
            # SafeGaurd: If User Edits Excel and Increases the Rows with Blanks, Stop Streaming
            if arduinoData[0].value == None:
                break
            
            # Add TimePoints
            self.analysisProtocol.data["timePoints"].append(arduinoData[0].value)
            # Add Biolectric Global Data to Dictionary in Sequential Order.
            for channelIndex in range(self.numChannels):
                channelVoltage = arduinoData[channelIndex+1].value
                self.analysisProtocol.data['Channel' + str(channelIndex+1)].append(float(channelVoltage or 0))  # Represent No Value (None) as 0
            
            pointNum += 1
            # When Ready, Send Data Off for Analysis
            while pointNum - dataFinger >= self.numTimePoints:
                self.analysisProtocol.analyzeData(dataFinger, plotStreamedData, predictionModel, actionControl = actionControl)
                dataFinger += self.moveDataFinger
        # At the End, Analyze Any Data Left
        if dataFinger < len(self.analysisProtocol.data["timePoints"]):
            self.analysisProtocol.analyzeData(dataFinger, plotStreamedData, predictionModel, actionControl = actionControl)

        # Finished Data Collection: Report Back to User
        print("\tDone Data Collecting from File: ", analyzeSheet.title)
        
    
    def getFeaturesFromExcel(self, excelSheet, Training_Data, Training_Labels, gestureClasses, labelMap):
        # Get Current Label fo the Signal
        currentLabel = excelSheet.title.split(" - ")[1]
        featureLabelIndexArray = np.where(gestureClasses == currentLabel.lower())[0]
        if labelMap:
            featureLabelIndexArray = [labelMap[featureLabelIndexArray]]
            if featureLabelIndexArray < 0:
                return [],[]
        # If Label Does Not Exist ... Warn User and Exit
        elif len(featureLabelIndexArray) == 0:
            print("Class Label", "'"+str(currentLabel)+"'", "Not Found")
            sys.exit()
        
        # If Excel Header Exists (Any Words), Skip Until You Find the Numerical Data
        for row in excelSheet.rows:
            if type(row[0].value) in [int, float]:
                dataStartRow = row[0].row
                break
        
        # Create Data Structure to Hold Results
        previousRowEmpty = False
        # Loop Through the Excel Worksheet to collect all the data
        for featureData in excelSheet.iter_rows(min_col=2+self.numChannels, min_row=dataStartRow, max_col=excelSheet.max_col, max_row=excelSheet.max_row):
            
            # If Features are Present
            if featureData[1+self.numChannels]:
                Training_Data.append(featureData[1+self.numChannels:])
                Training_Labels.append(featureLabelIndexArray[0])
            # If the Last Two Rows Were Empty, There are No More Features
            elif previousRowEmpty:
                break
            # If We Have an Empty Row, Check the Next
            else:
                previousRowEmpty = True         
                    
        print("\tCollected Training Data for:", excelSheet.title)
        return Training_Data, Training_Labels
    
    
    def getTrainingData(self, trainingDataExcelFolder, gestureClasses, featureLabels, labelMap, mode):
        """
        Parameters
        ----------
        trainingDataExcelFolder: The Folder with ONLY the Training Data Excel Files
        gestureClasses: A List of Possible Classes (Represented as Strings)
        mode: The Type of Program to Run
            'Train' -> Get Trainign Data and Labels
            'reAnalyze' -> ReAnalyze All Data and Overwrite Excel File
        """
        Training_Data = []; Training_Labels = []
            
        for excelFile in list(os.listdir(trainingDataExcelFolder)):
            if excelFile.endswith(".xlsx") and not excelFile.startswith("~"):
                # Get Full Path to the Excel File
                trainingExcelFile = trainingDataExcelFolder + excelFile
                print("\nLoading Excel File", trainingExcelFile)
                # Load the Excel File
                if mode == 'reAnalyze':
                    # If reAnalyzing, Properly Save/Edit the Files
                    saveExcelData = saveExcel(self.numChannels)
                    # Copy the Excel File tin Case Something Goes Wrong
                    os.makedirs(trainingDataExcelFolder + "File Copies/", exist_ok=True)
                    shutil.copy(trainingExcelFile, trainingDataExcelFolder + "File Copies/" + excelFile)
                    # Delete the Current Excel File
                    os.remove(trainingExcelFile)
                    # Load in the Copy
                    trainingExcelFile = trainingDataExcelFolder + "File Copies/" + excelFile
                WB = xl.load_workbook(trainingExcelFile, data_only=True, read_only=True)
                WB_worksheets = WB.worksheets
                # Loop Over Each Sheet in the File
                for excelSheet in WB_worksheets:
                    # Analyze the Data in the Sheet
                    self.analysisProtocol.resetGlobalVariables()
                    
                    # Get the Training Data/Label from the Sheet
                    if mode == 'Train':
                        # Get the Feature Label
                        currentLabel = excelSheet.title.split(" - ")[1]
                        gestureIndex = np.where(gestureClasses == currentLabel.lower())[0][0]
                        featureLabel = labelMap[gestureIndex]
                        # If No Label, We Dont Want to Analyze its Features
                        if featureLabel < 0:
                            continue
                        
                        # Analyze the Data in the Sheet
                        print("\n\tExtracting Features From", excelSheet.title, "With Label", featureLabel)
                        self.streamExcelData(trainingExcelFile, analyzeSheet = excelSheet) 
                        Training_Data.extend(self.analysisProtocol.featureList)
                        Training_Labels.extend(len(self.analysisProtocol.featureList)*[featureLabel])
                    elif mode == 'reAnalyze':
                        print("\n\tReanalyzing Excel Sheet:", excelSheet.title)
                        # Analyze the Data in the Sheet
                        self.streamExcelData(trainingExcelFile, analyzeSheet = excelSheet)
                        # Delete the Previous Analysis from Excel (Save Name/Info)
                        sheetName = excelSheet.title
                        currentGesture = sheetName.split(" - ")[1]
                        # Overwrite Excel Sheet with new Analysis
                        saveExcelData.saveData(self.analysisProtocol.data, self.analysisProtocol.featureList, featureLabels, trainingDataExcelFolder, excelFile, sheetName=sheetName, currentGesture=currentGesture)
        # If reAnalyzing, Remove Copied Files
        if mode == 'reAnalyze':
            shutil.rmtree(trainingDataExcelFolder + "File Copies/")
        # Return Training Data and Labels
        return Training_Data, Training_Labels


class saveExcel:
    def __init__(self, numChannels):
        # Input Parameters
        self.numChannels = numChannels  # The Number of Biolectric Channels
        
        # Specify OpenPyxl Asthetics
        self.openpyxlColors = {
            0: "F67280",
            1: "F8B195",
            2: "99B898",
            3: "45ADA8",
            4: "C06C84",
            5: "BC9E82",
            }

    def saveData(self, data, featureList, featureLabels, saveDataFolder, saveExcelName,
                     sheetName = "Trial 1 - No Gesture", currentGesture = "Data", WB=None):        
        # Create Output File Directory to Save Data: If None
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Path to File to Save
        excel_file = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excel_file):
            print("\tSaving the Data as New Excel Workbook")
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        # Overwrite Previous Excel Sheet; Replacing Sheetname that was Edited
        elif WB:
            print("\tOverWriting Excel File:", excel_file)
            WB_worksheet = WB.create_sheet(sheetName)
            print("\tSaving Sheet as", WB_worksheet.title, "\n")
        # Loading in Previous Excel File, Creating New Sheet, Editing Trial Number in SheetName
        else:
            print("\tExcel File Already Exists. Loading File")
            WB = xl.load_workbook(excel_file, read_only=False)
            currentSheets = WB.sheetnames
            # Get All Sheets with the Current Movement
            currentMovementSheets = []
            for sheet in currentSheets:
                movement = sheet.split(" - ")
                if currentGesture == movement[1]:
                    currentMovementSheets.append(sheet)
            # Get the Last Trial for this Hand Movement
            sheetName = max(currentMovementSheets, key=lambda x: int(re.findall(r'\d+', x.split(" - ")[0])[0]), default= "Trial 0 - " + currentGesture)
            # Edit SheetName
            sheetInfo = sheetName.split(" - ")
            currentTrial = re.findall(r'\d+', sheetInfo[0])[0]
            newTrial = sheetInfo[0].split(currentTrial)[0] + str(int(currentTrial)+1)
            sheetName = newTrial + " - " + sheetInfo[1]
            # Add Sheet
            WB_worksheet = WB.create_sheet(sheetName)
            print("\tSaving Sheet as", sheetName)
        
        numFeatures = len(featureList[0]) if featureList else 0
        channelHeader = ['Channel ' + str(channelNum) + " Data" for channelNum in range(1, 1+self.numChannels)]
        # Creater Header
        header = ["timePoints"]
        header.extend(channelHeader)
        header.extend(featureLabels)
        # Label First Row
        WB_worksheet.append(header)
        
        numFeaturesAdded = 0
        # Save Bioelectric Data to Worksheet (First 1 + numChannels Columns)
        for dataNum in range(len(data['timePoints'])):
            # Add TimePoints
            row = [data['timePoints'][dataNum]]
            # Add Channel Data
            for channelNum in range(1, self.numChannels+1):
                row.append(data['Channel'+str(channelNum)][dataNum])
            # Add Features if Availible
            if numFeaturesAdded != len(featureList) and dataNum%2 == 0:
                row.extend(featureList[numFeaturesAdded])
                numFeaturesAdded += 1
            WB_worksheet.append(row)
            
        """        
        # Add Feature Locs (Next Columns) and Then Features (Next Next Columns)
        startIndex = 2 # Start at Secon Row (1-Indexed) After the Header
        for groupNum in range(len(featureList)):
            rowIndex = startIndex
            peakColor = (groupNum-1)%(len(self.openpyxlColors))
            cellColor = self.openpyxlColors[peakColor]
            for featureValList in featureList[groupNum]:
                for featureVal in featureValList:
                    # Add Feature Location
                    WB_worksheet.cell(row=rowIndex, column=channelIndex + self.numChannels + 2).fill = PatternFill(fgColor=cellColor, fill_type = 'solid')
        """
        
        alignCenter = Alignment(horizontal='center', vertical='center', wrap_text=True)  
        # Center the Data in the Cells
        for rowInd in range(2, WB_worksheet.max_row + 1):
            for colInd in range(1, self.numChannels + numFeatures + 2):
                WB_worksheet.cell(row=rowInd, column=colInd).alignment = alignCenter
        # Increase Cell Width to Encompass All Data and to be Even
        for column_cells in WB_worksheet.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            WB_worksheet.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length
        # Header Style
        for colInd in range(1, self.numChannels + numFeatures + 2):
            WB_worksheet.cell(row=1, column=colInd).font = Font(color='00FF0000', italic=True, bold=True)
            WB_worksheet.cell(row=1, column=colInd).alignment = alignCenter
        
        # Save as New Excel File
        WB.save(excel_file)
        WB.close()
    
    def saveLabeledPoints(self, signalData, signalLabelsTrue, signalLabelsPredicted, saveDataFolder, saveExcelName, sheetName = "Signal Data and Labels"): 
        # Create Output File Directory to Save Data: If None
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Path to File to Save
        excel_file = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excel_file):
            print("\nSaving the Data as New Excel Workbook", "\n")
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        # Loading in Previous Excel File and Creating New Sheet
        else:
            print("Excel File Already Exists. Loading File")
            WB = xl.load_workbook(excel_file, read_only=False)
            WB_worksheet = WB.create_sheet(sheetName)
            print("Saving Sheet as", sheetName, "\n")
        
        header = ['Dimension ' + str(featureNum) for featureNum in range(1, 1+len(signalData[0]))]
        header.extend(['Signal Labels True', 'Signal Labels Predicted'])
        WB_worksheet.append(header)
        
        # Save Data to Worksheet
        for pointInd in range(len(signalData)):
            # Get labels
            row = [signalData[pointInd][featureNum] for featureNum in range(len(signalData[pointInd]))]
            row.extend([signalLabelsTrue[pointInd], signalLabelsPredicted[pointInd]])
            WB_worksheet.append(row)
        
        # Center the Data in the Cells
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)        
        for column_cells in WB_worksheet.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            WB_worksheet.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length
            
            for cell in column_cells:
                cell.alignment = align
        # Header Style
        for cell in WB_worksheet["1:1"]:
            cell.font = Font(color='00FF0000', italic=True, bold=True)
            
        # Save as New Excel File
        WB.save(excel_file)
        WB.close()
