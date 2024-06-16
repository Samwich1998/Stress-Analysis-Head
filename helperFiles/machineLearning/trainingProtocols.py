import os
import re
import scipy
import numpy as np
import pandas as pd
from natsort import natsorted
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Import Files for Machine Learning
from .modelControl.Models.pyTorch.modelArchitectures.emotionModelInterface.emotionModel.emotionModelHelpers.modelParameters import modelParameters
from .modelControl.modelSpecifications.compileModelInfo import compileModelInfo
from .featureAnalysis.featurePlotting import featurePlotting

# Import excel data interface
from ..dataAcquisitionAndAnalysis.excelProcessing.extractDataProtocols import extractData
from ..dataAcquisitionAndAnalysis.excelProcessing.saveDataProtocols import saveExcelData


class trainingProtocols(extractData):

    def __init__(self, biomarkerFeatureNames, streamingOrder, biomarkerFeatureOrder, trainingFolder, readData):
        super().__init__()
        # General parameters
        self.biomarkerFeatureOrder = biomarkerFeatureOrder
        self.biomarkerFeatureNames = biomarkerFeatureNames
        self.numberOfChannels = len(streamingOrder)
        self.trainingFolder = trainingFolder
        self.streamingOrder = streamingOrder
        self.readData = readData

        # Extract feature information
        self.featureNames = [item for sublist in self.biomarkerFeatureNames for item in sublist]

        # Initialize important classes
        self.analyzeFeatures = featurePlotting(self.trainingFolder + "dataAnalysis/", overwrite=False)
        self.compileModelInfo = compileModelInfo()
        self.modelParameters = modelParameters
        self.saveInputs = saveExcelData()

    def streamTrainingData(self, featureAverageWindows, plotTrainingData=False, reanalyzeData=False, metaTraining=False, reverseOrder=False):
        # Hold time series analysis of features.
        allRawFeatureIntervals, allRawFeatureIntervalTimes, allAlignedFeatureIntervals, allAlignedFeatureIntervalTimes = [], [], [], []
        # Hold features extraction information.
        allRawFeatureHolders, allRawFeatureTimesHolders, allAlignedFeatureHolder, allAlignedFeatureTimes = [], [], [], []
        # Hold survey information.
        subjectInformationQuestions, surveyAnswersList, surveyAnswerTimes, surveyQuestions = [], [], [], []
        # Hold experimental information.
        subjectOrder, experimentalOrder = [], []

        # ---------------------- Collect Training Data --------------------- #

        # For each file in the training folder.
        for excelFile in natsorted(os.listdir(self.trainingFolder), reverse=reverseOrder):
            if not excelFile.endswith(".xlsx") or excelFile.startswith(("~", ".")): continue

            # Extract the file details
            trainingExcelFile = self.trainingFolder + excelFile
            excelFileName = excelFile.split(".")[0]

            # If the file ends with saveFeatureFile_Appended, then the file is a feature file.
            if not excelFile.endswith(self.saveFeatureFile_Appended):
                excelFile = self.saveFeatureFolder + excelFile.split(".")[0] + self.saveFeatureFile_Appended
            savedFeaturesFile = self.trainingFolder + excelFile

            # ---------------- Stream in the Training File --------------- #
            print(f"\nCompiling training features from {excelFileName}")

            # If you want to and can use previously extracted features
            if not reanalyzeData and os.path.isfile(savedFeaturesFile):
                print("\tUsing the previously analyzed features.")
                self.analyzeFeatures.overwrite = False

                # Read in the saved features
                rawFeatureTimesHolder, rawFeatureHolder, _, experimentTimes, experimentNames, currentSurveyAnswerTimes, \
                    currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions \
                    = self.getFeatures(self.biomarkerFeatureOrder, savedFeaturesFile, self.biomarkerFeatureNames, surveyQuestions, subjectInformationQuestions)

                # Add the parameters to readData.
                self.readData.subjectInformationAnswers = currentSubjectInformationAnswers
                self.readData.subjectInformationQuestions = subjectInformationQuestions
                self.readData.rawFeatureTimesHolder = rawFeatureTimesHolder
                self.readData.surveyAnswerTimes = currentSurveyAnswerTimes
                self.readData.surveyAnswersList = currentSurveyAnswersList
                self.readData.rawFeatureHolder = rawFeatureHolder
                self.readData.experimentTimes = experimentTimes
                self.readData.experimentNames = experimentNames
                self.readData.surveyQuestions = surveyQuestions
            else:
                print("\tReanalyzing the feature file.")
                self.analyzeFeatures.overwrite = True

                # Read in the training file with the raw data,
                WB = load_workbook(trainingExcelFile, data_only=True, read_only=True)

                self.readData.resetGlobalVariables()
                # Extract and analyze the raw data.
                compiledRawData, experimentTimes, experimentNames, currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions \
                    = self.extractExperimentalData(WB.worksheets, self.numberOfChannels, surveyQuestions=surveyQuestions, finalSubjectInformationQuestions=subjectInformationQuestions)
                self.readData.streamExcelData(compiledRawData, experimentTimes, experimentNames, currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions, excelFileName)
                # Extract information from the streamed data
                rawFeatureTimesHolder = self.readData.rawFeatureTimesHolder.copy()  # dim: numBiomarkers, numTimePoints
                rawFeatureHolder = self.readData.rawFeatureHolder.copy()  # dim: numBiomarkers, numTimePoints, numBiomarkerFeatures

                # Plot the signals
                self.analyzeFeatures.plotRawData(self.readData, compiledRawData, currentSurveyAnswerTimes, experimentTimes, experimentNames, self.streamingOrder, folderName=excelFileName + "/rawSignals/")

                # Save the features to be analyzed in the future.
                self.saveInputs.saveRawFeatures(rawFeatureTimesHolder, rawFeatureHolder, self.biomarkerFeatureNames, self.biomarkerFeatureOrder, experimentTimes, experimentNames, currentSurveyAnswerTimes,
                                                currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions, trainingExcelFile)
            # rawFeatureHolder dim: numBiomarkers, numTimePoints, numBiomarkerFeatures
            # rawFeatureTimesHolder dim: numBiomarkers, numTimePoints

            # ----------------------- Safety Checks ----------------------- #

            # Assert consistency across training data.
            assert len(experimentTimes) == len(currentSurveyAnswerTimes), f"Mismatch in experiment and survey times: {experimentTimes}, {currentSurveyAnswerTimes}"
            assert len(self.biomarkerFeatureOrder) == len(rawFeatureHolder), "Incorrect number of channels"

            noFeaturesFound = False
            # Checkpoint: are there features in ALL categories
            for biomarkerInd in range(len(rawFeatureTimesHolder)):
                if not len(rawFeatureTimesHolder[biomarkerInd]) != 0:
                    noFeaturesFound = True
            if noFeaturesFound: print("No features found"); continue

            # -------------------- Prepare Raw Features -------------------- #

            # Add the parameters to readData.
            self.readData.rawFeatureTimesHolder = rawFeatureTimesHolder
            self.readData.rawFeatureHolder = rawFeatureHolder

            # Convert to numpy arrays. Note, the holder may be inhomogeneous.
            for biomarkerInd in range(len(rawFeatureHolder)):
                rawFeatureTimesHolder[biomarkerInd] = np.asarray(rawFeatureTimesHolder[biomarkerInd])
                rawFeatureHolder[biomarkerInd] = np.asarray(rawFeatureHolder[biomarkerInd])

            # ----------------- Extract Compiled Features ----------------- #

            # Average the features across a sliding window at each timePoint
            compiledFeatureHolders = self.readData.compileStaticFeatures(rawFeatureTimesHolder, rawFeatureHolder, featureAverageWindows)
            # compiledFeatureHolders dim: numBiomarkers, numTimePoints, numBiomarkerFeatures

            # Assert the compiled features are the same length as the raw features
            assert len(compiledFeatureHolders[0][0]) == len(rawFeatureHolder[0][0]), "Compiled features are not the same length as the raw features"
            assert len(compiledFeatureHolders[0]) == len(rawFeatureHolder[0]), "Compiled features are not the same length as the raw features"
            assert len(compiledFeatureHolders) == len(rawFeatureHolder), "Compiled features are not the same length as the raw features"

            # For each unique analysis with features.
            for analysis in self.readData.featureAnalysisList:
                for featureChannelInd in range(len(analysis.featureChannelIndices)):
                    biomarkerInd = analysis.featureChannelIndices[featureChannelInd]

                    # Add back the relevant feature information.
                    analysis.compiledFeatures[featureChannelInd] = compiledFeatureHolders[biomarkerInd]
                    analysis.rawFeatureTimes[featureChannelInd] = rawFeatureTimesHolder[biomarkerInd]
                    analysis.rawFeatures[featureChannelInd] = rawFeatureHolder[biomarkerInd]

            # ------------------ Align and Predict Labels ------------------ #

            # Align all the features.
            self.readData.alignFeatures()
            # Extract the aligned features
            alignedFeatureTimes = np.asarray(self.readData.alignedFeatureTimes)
            alignedFeatures = np.asarray(self.readData.alignedFeatures).T
            # alignedFeatures dim: numTimePoints, numFeatures
            # alignedFeatureTimes dim: numTimePoints

            # If there are no aligned features.
            if len(self.readData.alignedFeatureTimes) == 0:
                print("No aligned features found?!? Wierd... why did you waste my time!")
                raise ValueError

            # Finished analyzing the data
            self.readData.resetGlobalVariables()

            # ----------- Segment the Experimental Feature Signals ----------- #

            badExperimentalInds = []
            # For each experiment performed in the trial.
            for experimentInd in range(len(experimentTimes)):
                startExperimentTime, endExperimentTime = experimentTimes[experimentInd]
                startSurveyTime = currentSurveyAnswerTimes[experimentInd]

                # Calculate the raw feature intervals
                newRawFeatureIntervalTimes, newRawFeatureIntervals = self.organizeRawFeatureIntervals(startExperimentTime, startSurveyTime, rawFeatureTimesHolder, rawFeatureHolder)

                # Check the features.
                if newRawFeatureIntervalTimes is None:
                    badExperimentalInds.append(experimentInd)
                    continue

                # Calculate the aligned feature intervals
                modelFeatureTimeBuffer = self.modelParameters.getTimeWindows()[-1] + self.modelParameters.getShiftInfo(submodel='maxShift') - 5
                alignedFeatureIntervals, alignedFeatureIntervalTimes = self.readData.compileModelFeatures(alignedFeatureTimes, alignedFeatures, startSurveyTime - modelFeatureTimeBuffer, startSurveyTime)

                # Save the interval information
                allAlignedFeatureIntervalTimes.append(alignedFeatureIntervalTimes)
                allRawFeatureIntervalTimes.append(newRawFeatureIntervalTimes)
                allAlignedFeatureIntervals.append(alignedFeatureIntervals)
                experimentalOrder.append(experimentNames[experimentInd])
                allRawFeatureIntervals.append(newRawFeatureIntervals)

                if metaTraining:
                    subjectOrder.append(int(re.search(r'\d+', excelFileName).group()))
                else:
                    subjectOrder.append(" ".join(excelFileName.split(" ")[1:]))

            # Remove indices where no features were collected.
            for experimentInd in sorted(badExperimentalInds, reverse=True):
                currentSurveyAnswersList = np.delete(currentSurveyAnswersList, experimentInd, axis=0)
                del currentSurveyAnswerTimes[experimentInd]

            # -------------------- Plot the features ------------------- #

            if not metaTraining and False:
                # Get the predicted data
                alignedFeatureLabels = self.readData.predictLabels()
                _, allTrueLabels = self.compileModelInfo.extractFinalLabels(currentSurveyAnswersList, finalLabels=[])

                # Plot the predicted stress of each person if model provided
                for modelInd in range(len(self.compileModelInfo.predictionOrder)):
                    providedLabels = allTrueLabels[modelInd][0] * np.ones(len(alignedFeatureTimes))
                    self.analyzeFeatures.plotPredictedScores(alignedFeatureTimes, providedLabels, allTrueLabels[modelInd], currentSurveyAnswerTimes, experimentTimes, experimentNames,
                                                             predictionType=self.compileModelInfo.predictionOrder[modelInd], folderName=excelFileName + "/realTimePredictions/")
                    self.analyzeFeatures.plotPredictedScores(alignedFeatureTimes, alignedFeatureLabels[modelInd], allTrueLabels[modelInd], currentSurveyAnswerTimes, experimentTimes, experimentNames,
                                                             predictionType=self.compileModelInfo.predictionOrder[modelInd], folderName=excelFileName + "/realTimePredictions/")

            if plotTrainingData:

                startBiomarkerFeatureIndex = 0
                for biomarkerInd in range(len(rawFeatureHolder)):
                    rawFeatureTimes = rawFeatureTimesHolder[biomarkerInd]
                    rawFeatures = rawFeatureHolder[biomarkerInd]
                    # rawFeatures dim: numTimePoints, numBiomarkerFeatures
                    # rawFeatureTimes dim: numTimePoints

                    # Get the training data's aligned features information
                    endBiomarkerFeatureIndex = startBiomarkerFeatureIndex + len(self.biomarkerFeatureNames[biomarkerInd])
                    alignedFeatureSet = np.asarray(alignedFeatures)[:, startBiomarkerFeatureIndex:endBiomarkerFeatureIndex]
                    startBiomarkerFeatureIndex = endBiomarkerFeatureIndex

                    # Plot each biomarker's features from the training file.
                    self.analyzeFeatures.singleFeatureAnalysis(self.readData, rawFeatureTimes, rawFeatures, self.biomarkerFeatureNames[biomarkerInd], preAveragingSeconds=0, averageIntervalList=[30, 60, 90],
                                                               surveyCollectionTimes=currentSurveyAnswerTimes, experimentTimes=experimentTimes, experimentNames=experimentNames,
                                                               folderName=excelFileName + "/Feature Analysis/singleFeatureAnalysis - " + self.biomarkerFeatureOrder[biomarkerInd].upper() + "/")
                    self.analyzeFeatures.singleFeatureAnalysis(self.readData, alignedFeatureTimes, alignedFeatureSet, self.biomarkerFeatureNames[biomarkerInd], preAveragingSeconds=featureAverageWindows[biomarkerInd], averageIntervalList=[0],
                                                               surveyCollectionTimes=currentSurveyAnswerTimes, experimentTimes=experimentTimes, experimentNames=experimentNames,
                                                               folderName=excelFileName + "/Feature Analysis/alignedFeatureAnalysis - " + self.biomarkerFeatureOrder[biomarkerInd].upper() + "/")

            # ------------------ Organize Information ------------------ #

            # Set up the compilation variables
            allRawFeatureTimesHolders.append(rawFeatureTimesHolder)
            allRawFeatureHolders.append(rawFeatureHolder)

            # Store the aligned feature information
            allAlignedFeatureTimes.append(alignedFeatureTimes)
            allAlignedFeatureHolder.append(alignedFeatures)

            # Save the survey labels.
            surveyAnswersList.extend(currentSurveyAnswersList)
            surveyAnswerTimes.append(currentSurveyAnswerTimes)

            # -------------------------------------------------------------- #

        # ---------------------- Compile Training Data --------------------- #

        # Organize the final labels for the features
        featureLabelTypes, allFinalLabels = [], []
        if not metaTraining:
            featureLabelTypes, allFinalLabels = self.compileModelInfo.extractFinalLabels(surveyAnswersList, allFinalLabels)

        # ------------------------------------------------------------------ #

        # Ensure the proper data structure.
        allFinalLabels, surveyQuestions, surveyAnswersList = np.asarray(allFinalLabels), np.asarray(surveyQuestions), np.asarray(surveyAnswersList)
        print(f'surveyQuestions: {surveyQuestions}')

        assert len(allRawFeatureTimesHolders) == len(allRawFeatureHolders) == len(allAlignedFeatureTimes) == len(allAlignedFeatureHolder)
        assert len(allRawFeatureIntervals) == len(allRawFeatureIntervalTimes) == len(allAlignedFeatureIntervals) == len(allAlignedFeatureIntervalTimes)

        # Return Training Data and Labels
        return allRawFeatureTimesHolders, allRawFeatureHolders, allRawFeatureIntervals, allRawFeatureIntervalTimes, \
            allAlignedFeatureTimes, allAlignedFeatureHolder, allAlignedFeatureIntervals, allAlignedFeatureIntervalTimes, \
            subjectOrder, experimentalOrder, allFinalLabels, featureLabelTypes, surveyQuestions, surveyAnswersList, surveyAnswerTimes

    def organizeRawFeatureIntervals(self, startExperimentTime, startSurveyTime, rawFeatureTimesHolder, rawFeatureHolder):
        # Append a new data point.
        newRawFeatureIntervalTimes = []
        newRawFeatureIntervals = []

        # For each biomarker during that trial
        for biomarkerInd in range(len(rawFeatureHolder)):
            rawFeatureTimes = rawFeatureTimesHolder[biomarkerInd]
            rawFeatures = np.array(rawFeatureHolder[biomarkerInd]).T
            # rawFeatures dim: numTimePoints, numBiomarkerFeatures
            # rawFeatureTimes dim: numTimePoints

            # Calculate the raw feature intervals
            featureIntervals, featureIntervalTimes = self.readData.compileModelFeatures(rawFeatureTimes, rawFeatures, startExperimentTime, startSurveyTime)

            # If there are no features found
            if featureIntervals is None:
                newRawFeatureIntervalTimes, newRawFeatureIntervals = None, None
                break

            # Save raw and aligned interval information
            newRawFeatureIntervalTimes.extend(featureIntervalTimes)
            newRawFeatureIntervals.extend(featureIntervals)

        return newRawFeatureIntervalTimes, newRawFeatureIntervals

    def varyAnalysisParam(self, dataFile, featureAverageWindows, featureTimeWindows):
        print("\nLoading Excel File", dataFile)
        # Read in the training file with the raw data,
        WB = load_workbook(dataFile, data_only=True, read_only=True)

        allRawFeatureTimesHolders = []
        allRawFeatureHolders = []

        # Extract and analyze the raw data.
        compiledRawData, experimentTimes, experimentNames, currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions \
            = self.extractExperimentalData(WB.worksheets, self.numberOfChannels, surveyQuestions=[], finalSubjectInformationQuestions=[])

        # For each test parameter
        for featureTimeWindow in featureTimeWindows:
            print(featureTimeWindow)
            # Set the parameter in the analysis
            self.readData.setFeatureWindowEEG(featureTimeWindow)

            # Stream the data
            self.readData.streamExcelData(compiledRawData, experimentTimes, experimentNames, currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions, dataFile)
            # Extract information from the streamed data
            allRawFeatureTimesHolders.append(self.readData.rawFeatureTimesHolder.copy())
            allRawFeatureHolders.append(self.readData.rawFeatureHolder.copy())
            # Remove all previous information from this trial
            self.readData.resetGlobalVariables()

            # Assert consistency across training data
            assert len(self.biomarkerFeatureOrder) == len(allRawFeatureHolders[-1]), "Incorrect number of channels"

        biomarkerInd = self.biomarkerFeatureOrder.index('eeg')
        # For each EEG Feature
        for featureInd in range(len(self.biomarkerFeatureNames[biomarkerInd])):

            heatMap = []
            finalTimePoints = np.arange(max(allRawFeatureTimesHolders, key=lambda x: x[biomarkerInd][0])[biomarkerInd][0], min(allRawFeatureTimesHolders, key=lambda x: x[biomarkerInd][-1])[biomarkerInd][-1], 0.1)
            # For each feature-variation within a certain time window
            for trialInd in range(len(allRawFeatureHolders)):
                averageWindow = featureAverageWindows[biomarkerInd]
                rawFeatures = np.asarray(allRawFeatureHolders[trialInd][biomarkerInd])[:, featureInd:featureInd+1]
                rawFeatureTimes = np.asarray(allRawFeatureTimesHolders[trialInd][biomarkerInd])

                # Perform the feature averaging
                compiledFeatures = self.readData.averageFeatures_static(rawFeatureTimes, rawFeatures, averageWindow, startTimeInd=0)

                # Interpolate all the features within the same time-window
                featurePolynomial = scipy.interpolate.interp1d(rawFeatureTimes, compiledFeatures, kind='linear')
                finalFeatures = featurePolynomial(finalTimePoints)

                # Track the heatmap
                heatMap.append(finalFeatures)

            for finalFeatures in heatMap:
                plt.plot(finalTimePoints, finalFeatures, linewidth=1)

            plt.show()

            vMin = np.asarray(heatMap)[:, 100:].min()
            vMax = np.asarray(heatMap)[:, 100:].max()
            # Plot the heatmap
            ax = sns.heatmap(pd.DataFrame(heatMap, index=featureTimeWindows, columns=np.round(finalTimePoints, 2)), robust=True, vmin=vMin, vmax=vMax, cmap='icefire', yticklabels='auto')
            ax.set(title='Feature:' + self.biomarkerFeatureNames[biomarkerInd][featureInd])
            # Save the Figure
            sns.set(rc={'figure.figsize': (7, 9)})
            plt.xlabel("Time (Seconds)")
            plt.ylabel("Sliding Window Size")
            fig = ax.get_figure()
            fig.savefig(self.biomarkerFeatureNames[biomarkerInd][featureInd] + "Sliding Window.png", dpi=300)
            plt.show()

        return allRawFeatureTimesHolders, allRawFeatureHolders
